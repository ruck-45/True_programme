import openpyxl

def row_count(file_path,sheet_name):
    workbook = openpyxl.load_workbook(file_path)
    sheet = workbook.get_sheet_by_name(sheet_name)
    return sheet.max_row

def column_count(file_path,sheet_name):
    workbook = openpyxl.load_workbook(file_path)
    sheet = workbook.get_sheet_by_name(sheet_name)
    return sheet.max_column

def read_element(file_path,sheet_name,row_no,col_no):
    workbook = openpyxl.load_workbook(file_path)
    sheet = workbook.get_sheet_by_name(sheet_name)
    return sheet.cell(row_no,col_no).value

def write_elemnt(file_path,sheet_name,row_no,col_no,value):
    workbook = openpyxl.load_workbook(file_path)
    sheet = workbook.get_sheet_by_name(sheet_name)
    sheet.cell(row_no,col_no).value = value
    workbook.save(file_path)
